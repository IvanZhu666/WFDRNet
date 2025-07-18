import argparse
import os
import shutil
import warnings
import random

import torch
import torch.nn.functional as F
import numpy as np
from torch.utils.data import DataLoader
from tensorboardX import SummaryWriter
from openpyxl import load_workbook, Workbook

from constant import RESIZE_SHAPE, NORMALIZE_MEAN, NORMALIZE_STD, ALL_CATEGORY
from data.mvtec_dataset import MVTecDataset
from eval import evaluate_single_model
from model.WFDRNet_Model import WFDRNet
from model.losses import cosine_similarity_loss, focal_loss

warnings.filterwarnings("ignore")

all_categories_data = []

def seed_everything(seed=42):
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False

def append_to_xlsx(file_path, category, total_detail_best, all_data=None):
    file_exists = os.path.isfile(file_path)
    if file_exists:
        wb = load_workbook(filename=file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Category", "image_AUROC", "pixel_AUROC", "pixel_AP"])

    data_to_write = [
        category,
        round(float(total_detail_best["image_AUROC"]), 4),
        round(float(total_detail_best["pixel_AUROC"]), 4),
        round(float(total_detail_best["pixel_AP"]), 4)
    ]
    ws.append(data_to_write)

    if all_data is not None:
        all_data.append(total_detail_best)

    wb.save(file_path)

def append_average_to_xlsx(file_path, all_data):
    if not all_data:
        print("No data to calculate average.")
        return

    avg_data = {
        "Category": "Average",
        "image_AUROC": sum(item["image_AUROC"] for item in all_data) / len(all_data),
        "pixel_AUROC": sum(item["pixel_AUROC"] for item in all_data) / len(all_data),
        "pixel_AP": sum(item["pixel_AP"] for item in all_data) / len(all_data),
    }

    for key in avg_data:
        if key != "Category":
            avg_data[key] = round(float(avg_data[key]), 4)

    append_to_xlsx(file_path, avg_data["Category"], avg_data)


def create_dataloader(args, category, rotate_90=False, random_rotate=0):
    dataset = MVTecDataset(
        is_train=True,
        mvtec_dir=os.path.join(args.mvtec_path, category, "train/good/"),
        resize_shape=RESIZE_SHAPE,
        normalize_mean=NORMALIZE_MEAN,
        normalize_std=NORMALIZE_STD,
        dtd_dir=args.dtd_path,
        rotate_90=rotate_90,
        random_rotate=random_rotate,
    )
    dataloader = DataLoader(
        dataset,
        batch_size=args.bs,
        shuffle=True,
        num_workers=args.num_workers,
        drop_last=True,
    )
    return dataloader


def create_model_and_optimizers(args):
    model = WFDRNet().cuda()
    args_lr_patch_embedding = 0.4
    args_lr_attention = 0.001
    args_lr_mlp = 0.001
    args_lr_conv = 0.1

    patch_embedding_params = []
    attention_params = []
    mlp_params = []
    conv_params = []

    for name, param in model.fdb.named_parameters():
        if 'patch_embed' in name:
            patch_embedding_params.append(param)
        elif 'attn' in name:
            attention_params.append(param)
        elif 'encoder' in name:
            mlp_params.append(param)
        elif 'upconv' in name or 'conv' in name:
            conv_params.append(param)

    fdb_optimizer = torch.optim.SGD(
        [
            {"params": patch_embedding_params, "lr": args_lr_patch_embedding},
            {"params": attention_params, "lr": args_lr_attention},
            {"params": mlp_params, "lr": args_lr_mlp},
            {"params": conv_params, "lr": args_lr_conv}
        ],
        lr=args_lr_patch_embedding,
        momentum=0.9,
        weight_decay=1e-4,
        nesterov=False,
    )

    frb_optimizer = torch.optim.SGD(
        [{"params": model.frb.parameters(), "lr": 0.4}],
        lr=0.4,
        momentum=0.9,
        weight_decay=1e-4,
        nesterov=False,
    )

    amg_optimizer = torch.optim.Adam(
        model.amg.parameters(), lr=0.001, weight_decay=1e-5
    )
    amg_scheduler = torch.optim.lr_scheduler.ReduceLROnPlateau(
        amg_optimizer,
        mode='min',
        factor=0.9,
        patience=args.patience_steps,
        verbose=True,
        threshold=1e-4,
        threshold_mode='rel',
        cooldown=0,
        min_lr=1e-7,
        eps=1e-8
    )

    return model, fdb_optimizer, frb_optimizer, amg_optimizer, amg_scheduler


def train_single_step(
    model,
    img_aug,
    img_origin,
    mask,
    global_step,
    args,
    fdb_optimizer,
    frb_optimizer,
    amg_optimizer,
    amg_scheduler
):
    if global_step < args.first_steps:
        model.frb.train()
        model.fdb.train()
        model.amg.eval()
    else:
        model.frb.eval()
        model.fdb.eval()
        model.amg.train()

    predicted_mask, output_first, first_mask = model(img_aug, img_origin)

    mask = F.interpolate(mask, size=predicted_mask.size()[2:], mode="bilinear", align_corners=False)
    mask = torch.where(mask < 0.5, torch.zeros_like(mask), torch.ones_like(mask))

    cosine_loss_val = cosine_similarity_loss(output_first)
    focal_loss_val = focal_loss(predicted_mask, mask, gamma=args.gamma)
    l1_loss_val = F.l1_loss(predicted_mask, mask, reduction="mean")

    if global_step < args.first_steps:
        total_loss_val = cosine_loss_val
        total_loss_val.backward()
        frb_optimizer.step()
        fdb_optimizer.step()
    else:
        total_loss_val = focal_loss_val + l1_loss_val
        total_loss_val.backward()
        amg_optimizer.step()
        amg_scheduler.step(total_loss_val)

    return cosine_loss_val, focal_loss_val, l1_loss_val, total_loss_val


def train_model_for_category(args, category, rotate_90=False, random_rotate=0, all_data=None):
    if not os.path.exists(args.checkpoint_path):
        os.makedirs(args.checkpoint_path)
    if not os.path.exists(args.log_path):
        os.makedirs(args.log_path)

    run_name = f"{args.run_name_head}_{category}"
    log_dir = os.path.join(args.log_path, run_name)
    if os.path.exists(log_dir):
        shutil.rmtree(log_dir)

    visualizer = SummaryWriter(log_dir=log_dir)

    model, fdb_optimizer, frb_optimizer, amg_optimizer, amg_scheduler = create_model_and_optimizers(args)
    dataloader = create_dataloader(args, category, rotate_90, random_rotate)

    global_step = 0
    flag = True
    total_detail_best = {'image_AUROC': 0, 'pixel_AUROC': 0, 'pixel_AP': 0}

    for _ in range(9999999):
        for _, sample_batched in enumerate(dataloader):
            amg_optimizer.zero_grad()
            fdb_optimizer.zero_grad()
            frb_optimizer.zero_grad()

            img_origin = sample_batched["img_origin"].cuda()
            img_aug = sample_batched["img_aug"].cuda()
            mask = sample_batched["mask"].cuda()

            mask = F.interpolate(mask, size=(64, 64), mode="bilinear", align_corners=False)
            mask = torch.where(mask < 0.5, torch.zeros_like(mask), torch.ones_like(mask))

            cosine_loss_val, focal_loss_val, l1_loss_val, total_loss_val = train_single_step(
                model,
                img_aug,
                img_origin,
                mask,
                global_step,
                args,
                fdb_optimizer,
                frb_optimizer,
                amg_optimizer,
                amg_scheduler
            )
            global_step += 1

            visualizer.add_scalar("cosine_loss", cosine_loss_val, global_step)
            visualizer.add_scalar("focal_loss", focal_loss_val, global_step)
            visualizer.add_scalar("l1_loss", l1_loss_val, global_step)
            visualizer.add_scalar("total_loss", total_loss_val, global_step)

            if global_step % args.eval_per_steps == 0:
                total_detail = evaluate_single_model(args, category, model, visualizer, global_step)
                if global_step > args.first_steps:
                    score_new = total_detail["pixel_AUROC"] + total_detail["pixel_AP"] + total_detail["image_AUROC"]
                    score_old = total_detail_best["pixel_AUROC"] + total_detail_best["pixel_AP"] + total_detail_best["image_AUROC"]
                    if score_new > score_old:
                        ckpt_path = os.path.join(args.checkpoint_path, f"{args.run_name_head}_{category}.pckl")
                        torch.save(model.state_dict(), ckpt_path)
                        print("\n\n\n")
                        print(f"Best {category} model updated: global step {global_step}")
                        print("\n\n\n")
                        total_detail_best = total_detail

            if global_step % args.log_per_steps == 0:
                if global_step < args.first_steps:
                    print(f"Training step {global_step}, cos loss: {round(float(cosine_loss_val), 4)}")
                else:
                    print(f"Training step {global_step}, focal: {round(float(focal_loss_val), 4)}, l1: {round(float(l1_loss_val), 4)}")

            if global_step >= args.steps:
                flag = False
                print("\n\n\n")
                print(f"Final result of category {category}:")
                print(f"image_AUROC: {round(float(total_detail_best['image_AUROC']), 4)}")
                print(f"pixel_AUROC: {round(float(total_detail_best['pixel_AUROC']), 4)}")
                print(f"pixel_AP:  {round(float(total_detail_best['pixel_AP']), 4)}")
                print("\n\n\n")
                results_xlsx = os.path.join(args.checkpoint_path, f"results{args.patience_steps}.xlsx")
                append_to_xlsx(results_xlsx, category, total_detail_best, all_data)
                break
        if not flag:
            break


def run_experiment(args):
    if args.custom_training_category:
        no_rotation_category = args.no_rotation_category
        slight_rotation_category = args.slight_rotation_category
        rotation_category = args.rotation_category
        for category in (no_rotation_category + slight_rotation_category + rotation_category):
            assert category in ALL_CATEGORY, f"{category} not in the ALL_CATEGORY list"
    else:
        no_rotation_category = ["capsule", "metal_nut", "pill", "toothbrush", "transistor"]
        slight_rotation_category = ["wood", "zipper", "cable"]
        rotation_category = ["screw", "bottle", "grid", "hazelnut", "leather", "tile", "carpet"]

    for category in rotation_category:
        print(f"Training category (rotate_90=True): {category}")
        train_model_for_category(args, category, rotate_90=True, random_rotate=5, all_data=all_categories_data)

    for category in no_rotation_category:
        print(f"Training category (no rotation): {category}")
        train_model_for_category(args, category, rotate_90=False, random_rotate=0, all_data=all_categories_data)

    for category in slight_rotation_category:
        print(f"Training category (slight rotation): {category}")
        train_model_for_category(args, category, rotate_90=False, random_rotate=5, all_data=all_categories_data)

    results_xlsx = os.path.join(args.checkpoint_path, f"results{args.patience_steps}.xlsx")
    append_average_to_xlsx(results_xlsx, all_categories_data)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--gpu_id", type=int, default=0)
    parser.add_argument("--num_workers", type=int, default=16)
    parser.add_argument("--mvtec_path", type=str, default="../datasets/mvtec/")
    parser.add_argument("--dtd_path", type=str, default="../datasets/dtd/images/")
    parser.add_argument("--checkpoint_path", type=str, default="./saved_model/")
    parser.add_argument("--run_name_head", type=str, default="Best")
    parser.add_argument("--log_path", type=str, default="./saved_model/logs/")
    parser.add_argument("--bs", type=int, default=16)

    parser.add_argument("--steps", type=int, default=10000)
    parser.add_argument("--first_steps", type=int, default=2000)
    parser.add_argument("--eval_per_steps", type=int, default=500)
    parser.add_argument("--log_per_steps", type=int, default=50)
    parser.add_argument("--patience_steps", type=int, default=500)

    parser.add_argument("--gamma", type=float, default=4)
    parser.add_argument("--T", type=int, default=100)

    parser.add_argument("--custom_training_category", action="store_true", default=False)
    parser.add_argument("--no_rotation_category", nargs="*", type=str, default=[])
    parser.add_argument("--slight_rotation_category", nargs="*", type=str, default=[])
    parser.add_argument("--rotation_category", nargs="*", type=str, default=[])

    args = parser.parse_args()

    seed_everything(42)

    with torch.cuda.device(args.gpu_id):
        run_experiment(args)

if __name__ == "__main__":
    main()
